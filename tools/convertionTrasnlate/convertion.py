import os
import polib
import openpyxl
from openpyxl.styles import Alignment

def preprocess_po(src_po_file, temp_po_file):
    rows = []
    with open(src_po_file, 'r') as reader:
        for ctx in reader.readlines():
            if ctx.find('#~') == 0:
                continue
            rows.append(ctx)
    with open(temp_po_file, 'w') as writer:
        for line in rows:
            writer.write(line)


def po_to_xlsx(po_file, xlsx_file):
    """将po文件转换为xlsx文件"""
    preprocess_po(po_file, "temp.po")

    po = polib.pofile("temp.po")
    # po = polib.pofile(po_file)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # 设置列宽
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 50

    # 写入标题行
    worksheet['A1'] = 'msgid'
    worksheet['B1'] = 'msgstr'
    worksheet['C1'] = 'comments'

    # 写入数据
    for i, entry in enumerate(po, start=2):
        worksheet.cell(row=i, column=1, value=entry.msgid)
        worksheet.cell(row=i, column=2, value=entry.msgstr)
        worksheet.cell(row=i, column=3, value=entry.comment)

    # 居中对齐
    for row in range(1, len(po) + 2):
        for col in range(1, 4):
            worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(xlsx_file)

def xlsx_to_po(xlsx_file, po_file):
    """将xlsx文件转换为po文件"""
    workbook = openpyxl.load_workbook(xlsx_file)
    worksheet = workbook.active

    po = polib.POFile()

    for row in range(2, worksheet.max_row + 1):
        msgid = worksheet.cell(row=row, column=1).value
        msgstr = worksheet.cell(row=row, column=2).value
        comment = worksheet.cell(row=row, column=3).value
        # must process None or you'll get errors
        if msgid is None:
            continue
        if msgstr is None:
            msgstr = ""
        if comment is None:
            comment = ""
        # entry = polib.POEntry(
        #     msgid=worksheet.cell(row=row, column=1).value,
        #     msgstr=worksheet.cell(row=row, column=2).value,
        #     # comment=worksheet.cell(row=row, column=3).value
        # )
        entry = polib.POEntry(
            msgid=msgid,
            msgstr=msgstr,
            comment=comment
            # comment=worksheet.cell(row=row, column=3).value
        )
       # print(msgstr)
        po.append(entry)

    po.save(po_file)

if __name__ == '__main__':
    # po_to_xlsx('Orca-Flashforge_zh_CN_raw.po', 'zh_cn_target.xlsx')
    # xlsx_to_po('zh_cn_target.xlsx', 'zh_cn_source.po')
    # po_to_xlsx('Orca-Flashforge_en_raw.po', 'en_target.xlsx')
    xlsx_to_po('korean.xlsx', 'ko.po')
    #po_to_xlsx('tobe_translate/Orca-Flashforge_de.po', 'tobe_translate/de_target.xlsx')
    # xlsx_to_po('tobe_translate/de_target.xlsx', 'en_translated.po')

    #po_to_xlsx('tobe_translate/Orca-Flashforge_ko.po', 'tobe_translate/ko_target.xlsx')
    #po_to_xlsx('tobe_translate/Orca-Flashforge_es.po', 'tobe_translate/es_target.xlsx')
    #po_to_xlsx('tobe_translate/Orca-Flashforge_fr.po', 'tobe_translate/fr_target.xlsx')

    #po_to_xlsx('tobe_translate/Orca-Flashforge_ja.po', 'tobe_translate/ja_target.xlsx')
    #po_to_xlsx('tobe_translate/Orca-Flashforge_it.po', 'tobe_translate/it_target.xlsx')
