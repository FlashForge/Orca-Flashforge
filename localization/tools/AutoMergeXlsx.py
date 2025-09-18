import os
import shutil
import traceback
import openpyxl
import PoRW

def _getXlsxMsgStrMap(xlsxFilePath):
    escapeTable = { ord("\n"):"\\n", ord("\r"):"\\r", ord("\t"):"\\t",\
                    ord("\\"):"\\\\", ord("\""):"\\\""}
    msgMap = {}
    workbook = openpyxl.load_workbook(xlsxFilePath)
    worksheet = workbook.active
    hasMsgCtxt = worksheet.cell(row=1,column=1).value == "msgctxt"
    for row in range(2, worksheet.max_row + 1):
        if not hasMsgCtxt:
            msgctxt = None
            msgid = worksheet.cell(row=row, column=1).value
            msgstr = worksheet.cell(row=row, column=2).value
        else:
            msgctxt = worksheet.cell(row=row, column=1).value
            msgid = worksheet.cell(row=row, column=2).value
            msgstr = worksheet.cell(row=row, column=3).value
        if msgctxt is None:
            msgctxt = ""
        if msgid is None:
            continue
        if msgstr is None:
            msgstr = ""
        key = PoRW.getMsgKeyRaw(msgctxt.translate(escapeTable), msgid.translate(escapeTable))
        msgMap[key] = msgstr.translate(escapeTable)
    return msgMap

def _appendMsgStrLines(lines, msgStr, splitBySpace):
    maxLineLen = 79
    if len(msgStr) <= maxLineLen - 7 and msgStr.find("\\n") == -1:
        lines.append("msgstr \"%s\"\n" % msgStr)
        return
    lines.append("msgstr \"\"\n")
    while True:
        lfIdx = msgStr.find("\\n")
        if lfIdx != -1 and lfIdx < maxLineLen - 1:
            lines.append("\"" + msgStr[:lfIdx + 2] + "\"\n")
            msgStr = msgStr[lfIdx + 2:]
            continue
        if len(msgStr) <= maxLineLen:
            if len(msgStr) > 0:
                lines.append("\"" + msgStr + "\"\n")
            break
        if splitBySpace:
            for i in range(maxLineLen, -1, -1):
                if i == 0:
                    lines.append("\"" + msgStr + "\"\n")
                    break
                elif msgStr[i].isspace() or msgStr[i] == '-':
                    lines.append("\"" + msgStr[:i + 1] + "\"\n")
                    msgStr = msgStr[i + 1:]
                    break
        else:
            lines.append("\"" + msgStr[:maxLineLen] + "\"\n")
            msgStr = msgStr[maxLineLen:]

def _getNewMsgLines(msg, newMsgstr, splitBySpace):
    lines = []
    isMsgStr = None
    for line in msg.lines:
        lineTrimed = line.lstrip()
        if lineTrimed.startswith("msgstr") and lineTrimed[6].isspace():
            _appendMsgStrLines(lines, newMsgstr, splitBySpace)
            isMsgStr = True
        elif lineTrimed.startswith("msgstr[0]"):
            isMsgStr = True
        elif lineTrimed.startswith("msgstr[1]"):
            isMsgStr = True
        elif lineTrimed.startswith("msgctxt")\
          or lineTrimed.startswith("msgid")\
          or lineTrimed.startswith("msgid_plural"):
            lines.append(line)
            isMsgStr = False
        elif not isMsgStr:
            lines.append(line)
    return lines

def _mergeXlsx(xlsxFilePath, poFilePath, splitBySpace):
    xlsxMsgStrMap = _getXlsxMsgStrMap(xlsxFilePath)
    poMsgList, poInvalidMsgList = PoRW.readMsgList(poFilePath, False)
    if len(poInvalidMsgList) != 0:
        print("bad po file")
        return
    poMsgMap = {}
    for msg in poMsgList:
        poMsgMap[PoRW.getMsgKey(msg)] = msg
    for key in xlsxMsgStrMap:
        if not key in poMsgMap:
            print("msgid %s not found" % key)
            return
    poFile = open("temp.po", "w", encoding="utf-8")
    for i, msg in enumerate(poMsgList):
        msgstr = xlsxMsgStrMap.get(PoRW.getMsgKey(msg))
        if msgstr is not None and len(msgstr) > 0 and msg.msgStr != msgstr:
            poFile.writelines(_getNewMsgLines(msg, msgstr, splitBySpace))
        else:
            poFile.writelines(msg.lines)
        if i != len(poMsgList) - 1:
            poFile.write("\n")
    poFile.close()
    shutil.copy("temp.po", poFilePath)

if __name__ == '__main__':
    try:
        appDir = os.path.dirname(os.path.abspath(__file__))
        for lan in ["de", "en", "es", "fr", "ja", "ko", "lt"]:
            xlsxFilePath = os.path.join(appDir, "flashforge_%s.xlsx" % lan)
            poFilePath = os.path.join(appDir, "../flashforge", lan, "flashforge_%s.po" % lan)
            if os.path.exists(xlsxFilePath):
                print("merge %s" % xlsxFilePath);
                splitBySpace = lan in ["de", "en", "es", "fr", "ko", "lt"];
                _mergeXlsx(xlsxFilePath, poFilePath, splitBySpace)
            else:
                print("%s not found" % xlsxFilePath)
    except:
        traceback.print_exc()
    os.system("pause")
