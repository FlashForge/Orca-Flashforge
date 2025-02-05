import os
import traceback
import openpyxl
import polib
from openpyxl.styles import Alignment

def _getUnfinishedXlsx(poFileFath, xlsxFilePath):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 50

    worksheet['A1'] = 'msgctxt'
    worksheet['B1'] = 'msgid'
    worksheet['C1'] = 'msgstr'
    worksheet['D1'] = 'comments'

    po = polib.pofile(poFileFath)
    row = 2
    for entry in po:
        if not entry.fuzzy and len(entry.msgstr) == 0:
            worksheet.cell(row=row, column=1, value=entry.msgctxt)
            worksheet.cell(row=row, column=2, value=entry.msgid)
            worksheet.cell(row=row, column=3, value=entry.msgstr)
            worksheet.cell(row=row, column=4, value=entry.comment)
            row += 1

    alignment = Alignment(horizontal='center', vertical='center')
    for row in range(1, len(po) + 2):
        for col in range(1, 5):
            worksheet.cell(row=row, column=col).alignment = alignment

    workbook.save(xlsxFilePath)

if __name__ == '__main__':
    try:
        app_dir = os.path.dirname(os.path.abspath(__file__))
        for lan in ["de", "en", "es", "fr", "ja", "ko", "lt"]:
            poFileName = "flashforge_%s.po" % lan
            poFilePath = os.path.join(app_dir, "../flashforge", lan, poFileName)
            xlsxFilePath = os.path.join(app_dir, "flashforge_%s.xlsx" % lan)
            _getUnfinishedXlsx(poFilePath, xlsxFilePath)
    except:
        traceback.print_exc()
        os.system("pause")
