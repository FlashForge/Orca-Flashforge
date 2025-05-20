import os
import sys
import traceback
import openpyxl
import polib
from openpyxl.styles import Alignment

def convertPoToXlsx(poFilePath, xlsxFilePath):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 50

    worksheet['A1'] = 'msgctxt'
    worksheet['B1'] = 'msgid'
    worksheet['C1'] = 'msgstr'
    worksheet['D1'] = 'comments'

    po = polib.pofile(poFilePath)
    for i, entry in enumerate(po, start=2):
        if not entry.fuzzy:
            worksheet.cell(row=i, column=1, value=entry.msgctxt)
            worksheet.cell(row=i, column=2, value=entry.msgid)
            worksheet.cell(row=i, column=3, value=entry.msgstr)
            worksheet.cell(row=i, column=4, value=entry.comment)

    alignment = Alignment(horizontal='center', vertical='center')
    for row in range(1, len(po) + 2):
        for col in range(1, 5):
            worksheet.cell(row=row, column=col).alignment = alignment

    workbook.save(xlsxFilePath)
        
if __name__ == '__main__':
    try:
        poFilePath = sys.argv[1]
        xlsxFilePath = os.path.splitext(os.path.basename(poFilePath))[0] + ".xlsx";
        convertPoToXlsx(poFilePath, xlsxFilePath)
    except:
        traceback.print_exc()
        os.system("pause")
        