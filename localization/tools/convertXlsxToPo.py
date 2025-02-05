import os
import sys
import traceback
import openpyxl
import polib

def _convertXlsxToPo(xlsxFilePath, poFilePath):
    po = polib.POFile()
    workbook = openpyxl.load_workbook(xlsxFilePath)
    worksheet = workbook.active
    hasMsgCtxt = worksheet.cell(row=1,column=1).value == "msgctxt"
    for row in range(2, worksheet.max_row + 1):
        if hasMsgCtxt:
            msgctxt = worksheet.cell(row=row, column=1).value
            msgid = worksheet.cell(row=row, column=2).value
            msgstr = worksheet.cell(row=row, column=3).value
        else:
            msgctxt = None
            msgid = worksheet.cell(row=row, column=1).value
            msgstr = worksheet.cell(row=row, column=2).value
        if msgid is None:
            continue
        if msgstr is None:
            msgstr = ""
        if msgctxt is None:
            entry = polib.POEntry(msgid=str(msgid), msgstr=str(msgstr))
        else:
            entry = polib.POEntry(msgctxt=str(msgctxt), msgid=str(msgid), msgstr=str(msgstr))
        po.append(entry)
    po.save(poFilePath)
        
if __name__ == '__main__':
    try:
        xlsxFilePath = sys.argv[1]
        poFilePath = os.path.splitext(os.path.basename(xlsxFilePath))[0] + ".po";
        _convertXlsxToPo(xlsxFilePath, poFilePath)
    except:
        traceback.print_exc()
        os.system("pause")
        